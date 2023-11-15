import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# could be change when address changes
DIR = '/Users/xuanzhentai/Happy Distro Files/Happy Distro Excel/Test File/'


def addMissingColumn(df1, df2):
    # different excel
    missing = set(df1.columns) - set(df2.columns)
    # print(missing)
    # Wed
    for column in missing:
        df2[column] = pd.NA
    if missing:
        df2 = reorderHeaders(df1, df2, missing)
    return df2


def addTemplateMissing(df, nameList):
    missing = set(nameList) - set(df.columns)
    # print(missing)
    for column in missing:
        df[column] = pd.NA
    if missing:
        df = reorderHeadersOne(df, missing)
    return df, missing


def reorderHeadersOne(df, missing):
    df_column = df.columns.tolist()
    index = df_column.index('type.1') - 1
    newOrder = df_column[:index] + list(missing) + df_column[index:-2] + list(missing)
    # print(f"neworder: {newOrder}")
    df = df[newOrder]
    return df


def reorderHeaders(df1, df2, missing):  # change wed, df1 is sat, df2 is wed
    preElement = None
    # 获取df1的列名列表
    df2_columns = df2.columns.tolist()
    for column in df1.columns:
        if column in missing:
            # 找到preElement在df2_columns中的位置，然后在其后面插入column
            index = df2_columns.index(preElement) + 1
            newOrder = df2_columns[:index] + [column] + df2_columns[index:]
            # 更新df2_columns为新的列顺序
            df2_columns = newOrder
        # 更新preElement为当前列名
        preElement = column
    df2_columns = df2_columns[: -len(missing)]
    # 重新排列df2的列
    df2 = df2[df2_columns]
    return df2


def renameDuplicated(df):
    # same excel
    # 初始化一个字典来跟踪列名出现的次数
    col_counts = {}
    # 新的列名列表
    new_columns = []
    # 遍历原始列名
    for col in df.columns:
        # 如果列名已经出现过，增加计数
        if col in col_counts:
            col_counts[col] += 1
            # 为第二次出现的列名添加后缀
            new_col = f"{col}.1" if col_counts[col] == 2 else col
        else:
            # 如果是第一次出现，计数设为1
            col_counts[col] = 1
            new_col = col
        # 添加到新列名列表
        new_columns.append(new_col)
    # 更新 DataFrame 的列名
    df.columns = new_columns
    return df


def get_team_name(sales_person, team_dict):

    # Function to get team name for a salesperson
    # Handle cases where the salesperson's name might be formatted differently
    for key in team_dict.keys():
        if key in sales_person:
            return team_dict[key]
    return "Unknown"  # Return "Unknown" if no match is found


def mergeAction(dfSat_sheet, dfWed_sheet, dfAdd, dfSO, dfSales, leaderName, location, hide):
    """
    This method merge Excel files and find the same sku and add or minus the quantity.
    :param dfSat_sheet: dfSat sheet
    :param dfWed_sheet: dfWed sheet
    :param dfAdd: dfAdd, only one sheet
    :param dfSO: dfSo only one sheet
    :param leaderName: the name of leader
    :param location: CA or TX
    :param hide: integer, hide sku and flavor or not
    :return: Final DataFrame
    """
    merged_df = None
    leaderNameTX = leaderName[:-2]

    if dfSat_sheet is None:
        pass
    elif dfWed_sheet is None:
        pass
    else:
        pass
    # 合并dfSat和dfWed
    merged_df1 = pd.merge(dfSat_sheet, dfWed_sheet, on='sku', how='outer', suffixes=('_sat', '_wed'))

    # 筛选出以 'B' 或 'C' 开头的 sku，因为sku这个column有不属于sku的东西
    mask = merged_df1['sku'].str.startswith('B') | merged_df1['sku'].str.startswith('C')
    merged_df1 = merged_df1[mask]

    # access additional excel
    if location == "tx":
        merged_df = mergeAdditional(dfAdd, leaderNameTX, merged_df1, location)
    else:
        merged_df = mergeAdditional(dfAdd, leaderName, merged_df1, location)

    # deal with SO Excel file
    merged_df = dealSOFile(merged_df, dfSO, dfSales, location, leaderName, leaderNameTX)

    # 填充NaN值
    fillNAN(merged_df, location, leaderName, leaderNameTX)

    # 计算group总和
    merged_df[leaderName + '_total'] = (merged_df[leaderName + '_sat'] + merged_df[leaderName + '_wed']
                                        + merged_df[location + "_allo"]) - merged_df['quantity']
    # 提取所需的列
    if not hide:
        desired_cols = ["sku", "flavor_sat", leaderName + "_sat", leaderName + "_wed",
                        location + "_allo", 'quantity',  leaderName + '_total']
    else:
        desired_cols = [leaderName + "_sat", leaderName + "_wed", location + "_allo", 'quantity', leaderName + '_total']
    dfOut = merged_df[desired_cols].copy()
    # rename columns name
    dfOut.rename(columns={"flavor_sat": "flavor/口味", leaderName + "_sat": location + ' ' + leaderName + ' Sat/周六',
                          leaderName + "_wed": location + ' ' + leaderName + ' Wed/周三',
                          location + '_allo': location + ' ' + leaderName + ' Additional/额外分配',
                          leaderName + "_total": location + ' ' + leaderName + " Total/总数"},
                 inplace=True)
    return dfOut


def calculate_cartons(item_name, quantity):
    # Function to calculate cartons from bars
    # Regular expression to find the number of bars
    match = re.search(r'(\d+)\s+(bars|bar)', item_name, re.IGNORECASE)
    if match:
        num_bars_per_box = int(match.group(1))
        total_bars = num_bars_per_box * quantity  # Total number of bars
        num_cartons = total_bars / 200  # Assuming 200 bars per carton
        return round(num_cartons)  # Round to the nearest whole number
    return None  # Return the original quantity if no match is found


def dealSOFile(merged_df, dfSO, dfSales, location, leaderName, leaderNameTX):
    # Create a dictionary to map salesperson names to teams
    team_dict = pd.Series(dfSales.team.values, index=dfSales.sales).to_dict()
    # Apply the function to the sales report DataFrame
    dfSO['team'] = dfSO['sales person'].apply(lambda col: get_team_name(col, team_dict))
    dfSO['team'] = dfSO['team'].str.lower()

    mask_warehouse = dfSO['warehouse'].str.lower().str.startswith(location)
    # Check if input name is a substring of the Sales Person column
    mask_salesperson = (dfSO['team'] == leaderName) | (dfSO['team'] == leaderNameTX)

    # Apply the function and update 'Usage unit' where necessary
    mask_box = dfSO['usage unit'] == 'box'
    dfSO.loc[mask_box, 'quantity'] = dfSO.loc[mask_box].apply(
        lambda row: calculate_cartons(row['item name'], row['quantity']), axis=1)
    dfSO.loc[mask_box, 'usage unit'] = 'carton'

    # Filter for cartons
    mask_carton = dfSO['usage unit'] == 'carton'
    filtered_carton_df = dfSO[mask_warehouse & mask_carton & mask_salesperson]
    grouped_carton_df = filtered_carton_df.groupby('sku').agg({'quantity': 'sum'}).reset_index()

    # Combine the two dataframes
    return pd.merge(merged_df, grouped_carton_df, on='sku', how='left').fillna(0)


def seachTypes(merged_df, location, leaderName, leaderNameTX, week):
    # 首先找到所有以 leaderName 开头并以 '_sat' 结尾的列
    if location == 'ca':
        # 将列名、leaderName 和 week 转换为小写进行比较
        columns_to_sum = [col for col in merged_df.columns if col.lower().startswith(leaderName.lower()) and col.lower().endswith(week.lower())]
    else:
        columns_to_sum = [col for col in merged_df.columns if col.lower().startswith(leaderNameTX.lower()) and col.lower().endswith(week.lower())]
    #columns_to_sum = rename_duplicates(columns_to_sum, leaderName)
    # merged_df.columns = columns_to_sum
    for i, ele in enumerate(columns_to_sum):
        if ele.endswith('.1' + week):
            if location == 'ca':
                return columns_to_sum[:i]
            else:
                return columns_to_sum[i:]


def fillNAN(merged_df, location, leaderName, leaderNameTX):
    num_unique_skus = merged_df['sku'].nunique()
    columns_to_sum = seachTypes(merged_df, location, leaderName, leaderNameTX, '_sat')
    # 计算每个 sku 对应的指定列的总和，并将结果存储在名为 'Andy_sat' 的新列中
    merged_df[leaderName + '_sat'] = merged_df.groupby('sku')[columns_to_sum].transform('sum').sum(axis=1)
    merged_df.loc[0:num_unique_skus, leaderName + '_sat'] = merged_df.loc[0:num_unique_skus, leaderName + '_sat'].fillna(0)

    columns_to_sum = seachTypes(merged_df, location, leaderName, leaderNameTX, '_wed')
    merged_df[leaderName + '_wed'] = merged_df.groupby('sku')[columns_to_sum].transform('sum').sum(axis=1)
    merged_df.loc[0:num_unique_skus, leaderName + '_wed'] = merged_df.loc[0:num_unique_skus, leaderName + '_wed'].fillna(0)

    # Convert the column to float, turning any errors into NaN
    merged_df[location + "_allo"] = pd.to_numeric(merged_df[location + "_allo"], errors='coerce')
    merged_df.loc[0:num_unique_skus, location + '_allo'] = merged_df.loc[0:num_unique_skus, location + '_allo'].fillna(0)
    merged_df.loc[0:num_unique_skus, 'quantity'] = merged_df.loc[0:num_unique_skus, 'quantity'].fillna(0)


def mergeAdditional(dfAdd, leaderName, mergedDf, location):
    # 从dfAdd筛选出Manager为leaderName的行
    dfName = dfAdd[(dfAdd["manager"].str.lower() == leaderName)]
    dfName = dfName.groupby('sku').agg({location + '_allo': 'sum'}).reset_index()
    # 与已合并的merged_df_1进行合并
    return pd.merge(mergedDf, dfName[['sku', location + '_allo']], on='sku', how='left',
                    suffixes=('', '_' + leaderName.lower() + 'manager'))


def addNewSheet(fileName, dfOut, sheetName):
    """
    Add a new sheet base on the filename that is provided
    :param fileName: file name that want to add new sheet
    :param dfOut: Final DataFrame
    :param sheetName: name of the sheet
    :return: None
    """
    with pd.ExcelWriter(DIR + fileName + 'Sum.xlsx', engine='openpyxl', mode='a') as writer:
        dfOut.to_excel(writer, sheet_name=sheetName, index=False)
        print(sheetName + ' added successfully!')


def createNewFile(dfOut, sheetName, fileName):
    """
    Create a new file
    :param dfOut: Final DataFrame
    :param sheetName: Name of the sheet, like BC5000
    :param fileName: the file name program is going to create
    :return: file name, so when add to new sheet, it also required filename
    """
    # user input
    # user_input = input("Please enter filename you want to create: ")
    # create a new file
    with pd.ExcelWriter(DIR + fileName + 'Sum.xlsx', engine='openpyxl') as writer:
        dfOut.to_excel(writer, sheet_name=sheetName, index=False)
    print('Successfully create a new Excel file!')
    return fileName


def colorFileAndInsertCol(excel_path):
    # Load the Excel file
    book = load_workbook(excel_path)

    # Iterate over each sheet in the workbook
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]

        # Set fill colors
        fill_color1 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        fill_color2 = PatternFill(start_color='7efcdf', end_color='7eb5fc', fill_type='solid')  # Light blue
        fill_color4 = PatternFill(start_color='7efc8d', end_color='7efc8d', fill_type='solid')  # Green

        sheet['A1'].fill = fill_color4

        # Set B to G columns to one color
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            cell = sheet[f'{col}1']
            cell.fill = fill_color1

        # Set I to M columns to another color after inserting a new column
        for col in ['I', 'J', 'K', 'L', 'M']:
            cell = sheet[f'{col}1']
            cell.fill = fill_color2

    # Save the changes
    book.save(excel_path)
